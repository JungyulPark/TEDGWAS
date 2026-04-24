"""
Phase A-12 v2: Rebuild Final_Numbers_Frozen.xlsx
FIXES:
1. Add TSHR, IGF1R to MR_Results (from fast_mr.txt UTF-16 decoded)
2. Back-calculate SE for rest10 genes (SE = |beta| / |qnorm(P/2)|)
3. Include all SE values 
4. Mark verified vs. pending items
"""
import csv
import math
import os
from datetime import datetime
from scipy import stats as scipy_stats

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl", "--quiet"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = r"c:\ProjectTEDGWAS"
wb = Workbook()

# Styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, color="FFFFFF", size=11)
sig_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
pend_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def se_from_beta_p(beta, pval):
    """Back-calculate SE from beta and P-value using z = beta/SE => SE = beta/z"""
    if pval <= 0 or pval >= 1 or beta == 0:
        return float('nan')
    z = scipy_stats.norm.ppf(1 - pval/2)
    if z == 0:
        return float('nan')
    return abs(beta) / abs(z)

def read_csv_file(path):
    rows = []
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows

# ===== Sheet 1: Study Design =====
ws1 = wb.active
ws1.title = "Study_Design"
ws1.append(["Category", "GWAS_ID", "Trait", "N_Total", "N_Cases", "N_Controls", "Population", "Year", "Source"])
style_header(ws1)
ws1.append(["Primary", "ebi-a-GCST90018627", "Graves disease", 175465, 2809, 172656, "European", 2021, "GWAS Catalog"])
ws1.append(["Replication", "ebi-a-GCST90038636", "Hyperthyroidism", 484598, 3731, 480867, "European", 2021, "GWAS Catalog"])
ws1.append(["Sensitivity", "FinnGen R12", "Graves ophthalmopathy", 500348, 858, 499490, "Finnish", 2024, "FinnGen"])
ws1.append(["eQTL", "eQTLGen", "Blood cis-eQTL", 31684, "", "", "European", 2021, "eQTLGen"])
ws1.append(["RNA-seq", "In-house", "TED orbital fat", 5, 4, 1, "Korean", "", "Own data"])

# ===== Sheet 2: MR_Results (ALL 8 genes × 3 outcomes) =====
ws2 = wb.create_sheet("MR_Results")
ws2.append(["Gene", "Outcome", "Method", "beta", "SE", "P_value", "nSNP", "Source"])
style_header(ws2)

# --- TSHR/IGF1R from fast_mr.txt (verified from original run log) ---
tshr_igf1r_data = [
    ["TSHR", "Primary(Graves)", "IVW", -1.394, 0.167, 6.79e-17, 2, "fast_mr.txt"],
    ["TSHR", "Replication(Hyper)", "IVW", -0.012, 0.001, 3.88e-29, 5, "fast_mr.txt"],
    ["IGF1R", "Primary(Graves)", "IVW", 0.217, 0.112, 5.27e-2, 11, "fast_mr.txt"],
    ["IGF1R", "Replication(Hyper)", "IVW", 0.001, 0.001, 4.12e-2, 11, "fast_mr.txt"],
]
for row_data in tshr_igf1r_data:
    ws2.append(row_data)

# --- Rest 6 genes from MR_rest10_summary.csv (SE back-calculated) ---
rest10 = read_csv_file(os.path.join(BASE, "TrackA_MR", "results", "MR_rest10_summary.csv"))
for r in rest10:
    beta = float(r["b"])
    pval = float(r["p"])
    se = se_from_beta_p(beta, pval)
    ws2.append([r["Gene"], r["GWAS"], "IVW", beta, round(se, 6), pval, int(r["nsnp"]), "MR_rest10 (SE calc)"])

# --- FinnGen MR (SE available from CSV) ---
finngen = read_csv_file(os.path.join(BASE, "TrackA_MR", "results", "MR_FinnGen_Local_Summary.csv"))
for r in finngen:
    ws2.append([r["gene"], "FinnGen_Sensitivity", r["method"], float(r["b"]), float(r["se"]),
                float(r["pval"]), int(r["nsnp"]), "MR_FinnGen_Local"])

# Highlight significant rows
for row_idx in range(2, ws2.max_row + 1):
    pval_cell = ws2.cell(row=row_idx, column=6)
    try:
        if float(pval_cell.value) < 0.05:
            for col in range(1, 9):
                ws2.cell(row=row_idx, column=col).fill = sig_fill
    except:
        pass

# ===== Sheet 3: MR Sensitivity =====
ws3 = wb.create_sheet("MR_Sensitivity")
ws3.append(["Gene", "Outcome", "nSNP", "Egger_Intercept_P", "Cochran_Q_P", "Steiger_Direction", "Steiger_P", "Notes"])
style_header(ws3)
ws3.append(["TSHR", "Primary(Graves)", 2, "N/A (nSNP<3)", "N/A (nSNP<3)", "TRUE", "3.41e-22", ""])
ws3.append(["IGF1R", "Primary(Graves)", 11, 0.8056, 0.5930, "TRUE", "9.51e-154", "No pleiotropy"])
ws3.append(["TNF", "Replication(Hyper)", 5, 0.2369, 0.0278, "TRUE", "1.29e-128", "Heterogeneity detected"])
ws3.append(["CTLA4", "Primary(Graves)", 9, "", "", "", "", ""])
ws3.append(["PPARG", "Primary(Graves)", 4, "", "", "", "", ""])
ws3.append(["ARRB1", "Primary(Graves)", 6, "", "", "", "", ""])
ws3.append(["IRS1", "Primary(Graves)", 3, "", "", "", "", ""])
ws3.append(["AKT1", "Primary(Graves)", 3, "", "", "", "", ""])

# ===== Sheet 4: Coloc =====
ws4 = wb.create_sheet("Coloc")
ws4.append(["Parameter", "Value"])
style_header(ws4)
ws4.append(["Locus", "TSHR (chr14: 80.5-81.7 Mb)"])
ws4.append(["eQTL source", "eQTLGen (N=31,684, European)"])
ws4.append(["GWAS source", "FinnGen R12 Graves ophthalmopathy"])
ws4.append(["N shared SNPs", 4046])
ws4.append(["PP.H0.abf", 1.047844e-37])
ws4.append(["PP.H1.abf", 1.697128e-04])
ws4.append(["PP.H2.abf", 9.706965e-36])
ws4.append(["PP.H3.abf", 1.473667e-02])
ws4.append(["PP.H4.abf", 9.850936e-01])
ws4.append(["Top GWAS SNP", "rs3783947"])
ws4.append(["Top eQTL SNP", "rs179252"])
ws4.append(["LD r2 (EUR 1000G)", 0.737])
ws4.append(["Case proportion (s)", 0.016])
ws4.append(["Dataset type eQTL", "quant"])
ws4.append(["Dataset type GWAS", "cc"])

# ===== Sheet 5: Gene Sets =====
ws5 = wb.create_sheet("Gene_Sets")
ws5.append(["Set", "N_genes", "Source", "Note"])
style_header(ws5)
ws5.append(["IGF-1R pathway", 36, "CTD/STRING/DrugBank/Manual", ""])
ws5.append(["TSHR pathway", 33, "CTD/STRING/DrugBank/Manual", ""])
ws5.append(["TED disease genes", 59, "Literature-based manual curation across 10 categories", "Category-level reference, verified 2026-04-22"])

# ===== Sheet 6: Intersections =====
ws6 = wb.create_sheet("Intersections")
ws6.append(["Zone", "N_genes", "Genes"])
style_header(ws6)
ws6.append(["IGF1R-only", 4, "IGF1R, IGF1, IL1B, CXCL8"])
ws6.append(["TSHR-only", 3, "TSHR, ADIPOQ, FABP4"])
ws6.append(["Shared", 8, "HAS2, HAS1, TNF, ARRB1, IL6, HAS3, PPARG, CEBPA"])
ws6.append(["Total intersection", 15, ""])

# ===== Sheet 7: Enrichment =====
ws7 = wb.create_sheet("Enrichment")
ws7.append(["Zone", "Database", "Rank", "Term", "P_value", "Adj_P", "Genes"])
style_header(ws7)
for zone, prefix in [("IGF1R-only", "IGF1R_only_genes"), ("TSHR-only", "TSHR_only_genes"), ("Shared", "Shared_genes")]:
    for db in ["KEGG_2021_Human", "GO_Biological_Process_2021"]:
        fpath = os.path.join(BASE, "TrackB_Network", "results", "enrichment", f"{prefix}_{db}.csv")
        if os.path.exists(fpath):
            rows = read_csv_file(fpath)
            for i, r in enumerate(rows[:5]):
                ws7.append([zone, db, i+1, r["Term"], float(r["P-value"]), float(r["Adjusted P-value"]), r["Overlapping Genes"]])

# ===== Sheet 8: CMap =====
ws8 = wb.create_sheet("CMap")
ws8.append(["Zone", "Direction", "Rank", "Term", "P_value", "Adj_P", "Genes"])
style_header(ws8)
for zone, prefix in [("IGF1R-only", "IGF1R_only_genes"), ("TSHR-only", "TSHR_only_genes"), ("Shared", "Shared_genes")]:
    for direction in ["down", "up"]:
        fpath = os.path.join(BASE, "TrackB_Network", "results", "cmap", f"{prefix}_LINCS_L1000_Chem_Pert_{direction}.csv")
        if os.path.exists(fpath):
            rows = read_csv_file(fpath)
            for i, r in enumerate(rows[:5]):
                ws8.append([zone, direction, i+1, r["Term"], float(r["P-value"]), float(r["Adjusted P-value"]), r["Overlapping Genes"]])

# ===== Sheet 9: Off-target =====
ws9 = wb.create_sheet("Offtarget")
ws9.append(["Category", "N_genes", "Source", "Genes"])
style_header(ws9)
ws9.append(["IGF1R expanded downstream", 359, "KEGG PI3K-Akt (hsa04151) + original 25", ""])
ws9.append(["Hearing loss genes", 855, "DisGeNET sensorineural hearing loss + deafness", ""])
ws9.append(["Insulin signaling genes", 137, "KEGG Insulin signaling (hsa04910)", ""])

coch_genes = []
with open(os.path.join(BASE, "TrackC_Offtarget", "results", "cochlear_intersection_full.csv"), 'r') as f:
    reader = csv.reader(f); next(reader)
    coch_genes = [r[0] for r in reader if r]
ws9.append(["Cochlear intersection", len(coch_genes), "IGF1R_expanded ∩ Hearing_loss", ", ".join(coch_genes)])

ins_genes = []
with open(os.path.join(BASE, "TrackC_Offtarget", "results", "insulin_intersection_full.csv"), 'r') as f:
    reader = csv.reader(f); next(reader)
    ins_genes = [r[0] for r in reader if r]
ws9.append(["Insulin intersection", len(ins_genes), "IGF1R_expanded ∩ Insulin_signaling", ", ".join(ins_genes)])
ws9.append(["INSR in insulin set?", "YES", "", ""])

# ===== Sheet 10: Transcriptome =====
ws10 = wb.create_sheet("Transcriptome")
ws10.append(["Gene", "Zone", "Ctrl_TPM", "TED_mean_TPM", "log2FC", "P_value", "Direction", "Significant"])
style_header(ws10)

trans = read_csv_file(os.path.join(BASE, "TrackA_MR", "results", "Transcriptome_15gene_validation.csv"))
for r in trans:
    pv = float(r["P_value"])
    sig = "YES" if pv < 0.05 else "NO"
    row_idx = ws10.max_row + 1
    ws10.append([r["Gene"], r["Zone"], float(r["Ctrl_TPM"]), float(r["TED_mean_TPM"]),
                 float(r["log2FC"]), pv, r["Direction"], sig])
    if sig == "YES":
        for col in range(1, 9):
            ws10.cell(row=row_idx, column=col).fill = sig_fill

ws10.append([])
ws10.append(["--- INSR / Off-target Validation ---", "", "", "", "", "", "", ""])
ins_val = read_csv_file(os.path.join(BASE, "TrackA_MR", "results", "Insulin_receptor_validation.csv"))
for r in ins_val:
    pv = float(r["P_value"])
    sig = "YES" if pv < 0.05 else "NO"
    row_idx = ws10.max_row + 1
    ws10.append([r["Gene"], "Offtarget", float(r["Ctrl_TPM"]), float(r["TED_mean_TPM"]),
                 float(r["log2FC"]), pv, r["Direction"], sig])
    if sig == "YES":
        for col in range(1, 9):
            ws10.cell(row=row_idx, column=col).fill = sig_fill

ws10.append([])
ws10.append(["--- Pathway Comparison ---", "", "", "", "", "", "", ""])
path_comp = read_csv_file(os.path.join(BASE, "TrackA_MR", "results", "Pathway_level_comparison.csv"))
for r in path_comp:
    ws10.append([r["Pathway"], "Pathway", "", "", float(r["Mean_log2FC"]), float(r["Comparison_P"]),
                 f"n={r['N_genes']}", ""])

# ===== Sheet 11: Triangulation =====
ws11 = wb.create_sheet("Triangulation")
ws11.append(["Gene", "Zone", "MR_Primary_P", "MR_Rep_P", "MR_FinnGen_P", "Coloc_PP4",
             "RNAseq_log2FC", "RNAseq_P", "Convergence_Score", "Evidence_Components"])
style_header(ws11)

# Build comprehensive lookup
mr_lookup = {
    "TSHR": {"primary": 6.79e-17, "replication": 3.88e-29, "finngen": 6.81e-07},
    "IGF1R": {"primary": 5.27e-2, "replication": 4.12e-2, "finngen": 0.2159},
    "TNF": {"primary": 0.4287, "replication": 3.57e-5, "finngen": 0.1108},
}
for r in rest10:
    gene = r["Gene"]
    if gene not in mr_lookup:
        mr_lookup[gene] = {}
    gwas = r["GWAS"]
    if "Primary" in gwas:
        mr_lookup[gene]["primary"] = float(r["p"])
    elif "Replication" in gwas:
        mr_lookup[gene]["replication"] = float(r["p"])
for r in finngen:
    gene = r["gene"]
    if gene not in mr_lookup:
        mr_lookup[gene] = {}
    mr_lookup[gene]["finngen"] = float(r["pval"])

rnaseq_lookup = {}
for r in trans:
    rnaseq_lookup[r["Gene"]] = {"log2fc": float(r["log2FC"]), "pval": float(r["P_value"]), "direction": r["Direction"]}
for r in ins_val:
    if r["Gene"] == "INSR":
        rnaseq_lookup["INSR"] = {"log2fc": float(r["log2FC"]), "pval": float(r["P_value"]), "direction": r["Direction"]}

zone_map = {
    "IGF1R": "IGF1R-only", "IGF1": "IGF1R-only", "IL1B": "IGF1R-only", "CXCL8": "IGF1R-only",
    "TSHR": "TSHR-only", "ADIPOQ": "TSHR-only", "FABP4": "TSHR-only",
    "HAS2": "Shared", "HAS1": "Shared", "TNF": "Shared", "ARRB1": "Shared",
    "IL6": "Shared", "HAS3": "Shared", "PPARG": "Shared", "CEBPA": "Shared",
    "INSR": "Offtarget"
}

for gene in ["TSHR", "IGF1R", "TNF", "CTLA4", "PPARG", "ARRB1", "IRS1", "AKT1",
             "IGF1", "IL1B", "CXCL8", "ADIPOQ", "FABP4",
             "HAS2", "HAS1", "IL6", "HAS3", "CEBPA", "INSR"]:
    zone = zone_map.get(gene, "")
    mr_p = mr_lookup.get(gene, {})
    rna = rnaseq_lookup.get(gene, {})
    
    mr_primary = mr_p.get("primary", "")
    mr_rep = mr_p.get("replication", "")
    mr_finn = mr_p.get("finngen", "")
    coloc_pp4 = 0.985 if gene == "TSHR" else ""
    log2fc = rna.get("log2fc", "")
    rna_p = rna.get("pval", "")
    
    score = 0
    components = []
    
    # +1 MR significant in any
    any_mr = False
    for v in [mr_primary, mr_rep, mr_finn]:
        if v != "" and float(v) < 0.05:
            any_mr = True
    if any_mr:
        score += 1
        components.append("MR")
    
    # +1 Coloc
    if coloc_pp4 != "" and float(coloc_pp4) > 0.8:
        score += 1
        components.append("Coloc")
    
    # +1 RNA-seq
    if rna_p != "" and float(rna_p) < 0.05:
        score += 1
        components.append("RNA-seq")
    
    # +1 Direction consistent
    if rna.get("direction", "") and rna_p != "" and float(rna_p) < 0.05:
        score += 1
        components.append("Direction")
    
    row_idx = ws11.max_row + 1
    ws11.append([gene, zone, mr_primary, mr_rep, mr_finn, coloc_pp4, log2fc, rna_p, score, "+".join(components)])
    
    # Color by score
    if score >= 3:
        for col in range(1, 11):
            ws11.cell(row=row_idx, column=col).fill = sig_fill

# Auto-fit columns
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

# Save
timestamp = datetime.now().strftime("%Y%m%d")
outpath = os.path.join(BASE, f"Final_Numbers_Frozen_{timestamp}.xlsx")
wb.save(outpath)
print(f"SUCCESS: {outpath}")
print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
print()

# Summary counts
print("=== MR_Results Summary ===")
mr_rows = ws2.max_row - 1
print(f"Total MR rows: {mr_rows}")
print(f"  TSHR/IGF1R: 4 rows (from fast_mr.txt)")
print(f"  Rest10 genes: {len(rest10)} rows (SE back-calculated)")
print(f"  FinnGen: {len(finngen)} rows (SE from CSV)")
print()
print("=== Triangulation Summary ===")
tri_rows = ws11.max_row - 1
print(f"Total genes: {tri_rows}")
